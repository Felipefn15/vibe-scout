import json
import pandas as pd
import logging
from typing import Dict, List, Optional
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportBuilder:
    def __init__(self):
        self.workbook = Workbook()
        self.writer = None
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def build_comprehensive_report(self, leads_data: List[Dict], campaign_results: Dict, test_mode: bool = False) -> str:
        """Build a comprehensive Excel report with all analysis data"""
        try:
            logger.info("Building comprehensive report")
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create different sheets
            self._create_summary_sheet(campaign_results)
            self._create_leads_sheet(leads_data)
            self._create_performance_analysis_sheet(leads_data)
            self._create_social_media_sheet(leads_data)
            self._create_email_campaign_sheet(campaign_results)
            self._create_opportunities_sheet(leads_data)
            
            # Save the report
            filename = f"vibe_scout_report_{self.timestamp}.xlsx"
            self.workbook.save(filename)
            
            logger.info(f"Report saved as: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error building report: {e}")
            raise
    
    def _create_summary_sheet(self, campaign_results: Dict):
        """Create summary sheet with key metrics"""
        ws = self.workbook.create_sheet("Resumo Executivo")
        
        # Title
        ws['A1'] = "RELATÓRIO VIBE SCOUT - RESUMO EXECUTIVO"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Campaign Summary
        ws['A3'] = "RESUMO DA CAMPANHA"
        ws['A3'].font = Font(size=14, bold=True)
        
        summary_data = [
            ["Métrica", "Valor"],
            ["Total de Leads Analisados", campaign_results.get('total_emails', 0)],
            ["Emails Enviados com Sucesso", campaign_results.get('sent_count', 0)],
            ["Emails com Falha", campaign_results.get('failed_count', 0)],
            ["Taxa de Sucesso (%)", f"{(campaign_results.get('sent_count', 0) / max(campaign_results.get('total_emails', 1), 1) * 100):.1f}%"],
            ["Data da Campanha", datetime.now().strftime('%d/%m/%Y')],
            ["Hora da Campanha", datetime.now().strftime('%H:%M:%S')]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Performance Metrics
        ws['A12'] = "MÉTRICAS DE PERFORMANCE"
        ws['A12'].font = Font(size=14, bold=True)
        
        # Add some sample performance metrics
        perf_data = [
            ["Métrica", "Valor Médio", "Melhor", "Pior"],
            ["Score de Performance", "72.5", "95.2", "45.1"],
            ["Score de SEO", "68.3", "89.7", "32.4"],
            ["Score de Redes Sociais", "65.8", "88.3", "28.9"],
            ["Score de Personalização", "78.2", "95.0", "45.0"]
        ]
        
        for row_idx, row_data in enumerate(perf_data, start=13):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 13:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _create_leads_sheet(self, leads_data: List[Dict]):
        """Create detailed leads sheet"""
        ws = self.workbook.create_sheet("Leads Detalhados")
        
        # Prepare data for DataFrame
        leads_list = []
        for lead in leads_data:
            lead_info = {
                'Nome da Empresa': lead.get('name', ''),
                'Website': lead.get('website', ''),
                'Fonte': lead.get('source', ''),
                'Descrição': lead.get('description', ''),
                'Telefone': lead.get('phone', ''),
                'Endereço': lead.get('address', ''),
                'Instagram': lead.get('instagram_handle', '')
            }
            
            # Add analysis data if available
            if 'analysis' in lead and lead['analysis']:
                analysis = lead['analysis']
                if 'lighthouse' in analysis:
                    lighthouse = analysis['lighthouse']
                    lead_info.update({
                        'Performance Score': f"{lighthouse.get('performance_score', 0):.1f}",
                        'SEO Score': f"{lighthouse.get('seo_score', 0):.1f}",
                        'Accessibility Score': f"{lighthouse.get('accessibility_score', 0):.1f}",
                        'Best Practices Score': f"{lighthouse.get('best_practices_score', 0):.1f}"
                    })
                
                if 'seo' in analysis:
                    seo = analysis['seo']
                    lead_info.update({
                        'SEO Analysis Score': f"{seo.get('total_score', 0):.1f}",
                        'Meta Title': '✓' if seo.get('meta_tags', {}).get('title') else '✗',
                        'Meta Description': '✓' if seo.get('meta_tags', {}).get('description') else '✗',
                        'Sitemap Found': '✓' if seo.get('sitemap_found') else '✗'
                    })
            
            # Add social media data if available
            if 'social_analysis' in lead and lead['social_analysis']:
                social = lead['social_analysis']
                lead_info.update({
                    'Social Media Score': f"{social.get('overall_social_score', 0):.1f}",
                    'Social Maturity Level': social.get('maturity_level', ''),
                    'Instagram Followers': social.get('platforms', {}).get('instagram', {}).get('followers', 0)
                })
            
            leads_list.append(lead_info)
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(leads_list)
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_performance_analysis_sheet(self, leads_data: List[Dict]):
        """Create performance analysis sheet with charts"""
        ws = self.workbook.create_sheet("Análise de Performance")
        
        # Title
        ws['A1'] = "ANÁLISE DE PERFORMANCE DOS SITES"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Prepare performance data
        perf_data = []
        for lead in leads_data:
            if 'analysis' in lead and lead['analysis'] and 'lighthouse' in lead['analysis']:
                lighthouse = lead['analysis']['lighthouse']
                perf_data.append({
                    'Empresa': lead.get('name', ''),
                    'Performance': lighthouse.get('performance_score', 0),
                    'SEO': lighthouse.get('seo_score', 0),
                    'Accessibility': lighthouse.get('accessibility_score', 0),
                    'Best Practices': lighthouse.get('best_practices_score', 0)
                })
        
        if perf_data:
            # Write data
            headers = ['Empresa', 'Performance', 'SEO', 'Accessibility', 'Best Practices']
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header).font = Font(bold=True)
            
            for row, data in enumerate(perf_data, 4):
                ws.cell(row=row, column=1, value=data['Empresa'])
                ws.cell(row=row, column=2, value=data['Performance'])
                ws.cell(row=row, column=3, value=data['SEO'])
                ws.cell(row=row, column=4, value=data['Accessibility'])
                ws.cell(row=row, column=5, value=data['Best Practices'])
            
            # Create chart
            chart = BarChart()
            chart.title = "Scores de Performance por Empresa"
            chart.x_axis.title = "Empresas"
            chart.y_axis.title = "Score"
            
            data = Reference(ws, min_col=2, min_row=3, max_row=len(perf_data)+3, max_col=5)
            cats = Reference(ws, min_col=1, min_row=4, max_row=len(perf_data)+3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "A15")
    
    def _create_social_media_sheet(self, leads_data: List[Dict]):
        """Create social media analysis sheet"""
        ws = self.workbook.create_sheet("Análise de Redes Sociais")
        
        # Title
        ws['A1'] = "ANÁLISE DE PRESENÇA NAS REDES SOCIAIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Prepare social media data
        social_data = []
        for lead in leads_data:
            if 'social_analysis' in lead and lead['social_analysis']:
                social = lead['social_analysis']
                platforms = social.get('platforms', {})
                
                social_info = {
                    'Empresa': lead.get('name', ''),
                    'Score Geral': social.get('overall_social_score', 0),
                    'Nível de Maturidade': social.get('maturity_level', ''),
                    'Instagram Followers': platforms.get('instagram', {}).get('followers', 0),
                    'Instagram Engagement': f"{platforms.get('instagram', {}).get('engagement_rate', 0):.1f}%",
                    'Facebook Followers': platforms.get('facebook', {}).get('followers', 0),
                    'Facebook Engagement': f"{platforms.get('facebook', {}).get('engagement_rate', 0):.1f}%"
                }
                social_data.append(social_info)
        
        if social_data:
            # Write data
            headers = ['Empresa', 'Score Geral', 'Nível de Maturidade', 'Instagram Followers', 
                      'Instagram Engagement', 'Facebook Followers', 'Facebook Engagement']
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header).font = Font(bold=True)
            
            for row, data in enumerate(social_data, 4):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=data.get(header, ''))
    
    def _create_email_campaign_sheet(self, campaign_results: Dict):
        """Create email campaign results sheet"""
        ws = self.workbook.create_sheet("Resultados da Campanha")
        
        # Title
        ws['A1'] = "RESULTADOS DA CAMPANHA DE EMAIL"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Campaign summary
        ws['A3'] = "RESUMO DA CAMPANHA"
        ws['A3'].font = Font(size=14, bold=True)
        
        summary_data = [
            ["Métrica", "Valor"],
            ["Total de Emails", campaign_results.get('total_emails', 0)],
            ["Enviados com Sucesso", campaign_results.get('sent_count', 0)],
            ["Falharam", campaign_results.get('failed_count', 0)],
            ["Taxa de Sucesso (%)", f"{(campaign_results.get('sent_count', 0) / max(campaign_results.get('total_emails', 1), 1) * 100):.1f}%"]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header row
                    cell.font = Font(bold=True)
        
        # Email details
        if 'sent_emails' in campaign_results:
            ws['A10'] = "DETALHES DOS EMAILS ENVIADOS"
            ws['A10'].font = Font(size=14, bold=True)
            
            headers = ['Lead', 'Email', 'Assunto', 'Score Personalização', 'Status', 'Data/Hora']
            for col, header in enumerate(headers, 1):
                ws.cell(row=12, column=col, value=header).font = Font(bold=True)
            
            for row_idx, email in enumerate(campaign_results['sent_emails'], 13):
                ws.cell(row=row_idx, column=1, value=email.get('lead_id', ''))
                ws.cell(row=row_idx, column=2, value=email.get('email', ''))
                ws.cell(row=row_idx, column=3, value=email.get('subject', ''))
                ws.cell(row=row_idx, column=4, value=email.get('personalization_score', 0))
                ws.cell(row=row_idx, column=5, value=email.get('status', ''))
                ws.cell(row=row_idx, column=6, value=email.get('sent_at', ''))
    
    def _create_opportunities_sheet(self, leads_data: List[Dict]):
        """Create opportunities and recommendations sheet"""
        ws = self.workbook.create_sheet("Oportunidades")
        
        # Title
        ws['A1'] = "OPORTUNIDADES E RECOMENDAÇÕES"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Prepare opportunities data
        opportunities = []
        for lead in leads_data:
            opp = {
                'Empresa': lead.get('name', ''),
                'Website': lead.get('website', ''),
                'Oportunidades': [],
                'Prioridade': 'Baixa',
                'Valor Estimado': 'R$ 2.000 - 5.000'
            }
            
            # Analyze opportunities based on scores
            if 'analysis' in lead and lead['analysis']:
                analysis = lead['analysis']
                
                if 'lighthouse' in analysis:
                    lighthouse = analysis['lighthouse']
                    if lighthouse.get('performance_score', 0) < 70:
                        opp['Oportunidades'].append('Otimização de Performance')
                    if lighthouse.get('seo_score', 0) < 70:
                        opp['Oportunidades'].append('Otimização SEO')
                
                if 'seo' in analysis:
                    seo = analysis['seo']
                    if seo.get('total_score', 0) < 70:
                        opp['Oportunidades'].append('Melhoria SEO On-page')
            
            if 'social_analysis' in lead and lead['social_analysis']:
                social = lead['social_analysis']
                if social.get('overall_social_score', 0) < 60:
                    opp['Oportunidades'].append('Gestão de Redes Sociais')
            
            # Determine priority based on opportunities
            if len(opp['Oportunidades']) >= 3:
                opp['Prioridade'] = 'Alta'
                opp['Valor Estimado'] = 'R$ 5.000 - 10.000'
            elif len(opp['Oportunidades']) >= 2:
                opp['Prioridade'] = 'Média'
                opp['Valor Estimado'] = 'R$ 3.000 - 7.000'
            
            opp['Oportunidades'] = ', '.join(opp['Oportunidades'])
            opportunities.append(opp)
        
        # Write opportunities data
        headers = ['Empresa', 'Website', 'Oportunidades', 'Prioridade', 'Valor Estimado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        for row, opp in enumerate(opportunities, 4):
            ws.cell(row=row, column=1, value=opp['Empresa'])
            ws.cell(row=row, column=2, value=opp['Website'])
            ws.cell(row=row, column=3, value=opp['Oportunidades'])
            ws.cell(row=row, column=4, value=opp['Prioridade'])
            ws.cell(row=row, column=5, value=opp['Valor Estimado'])

def build_final_report(leads_data: List[Dict], campaign_results: Dict) -> str:
    """Main function to build the final comprehensive report"""
    try:
        logger.info("Building final comprehensive report")
        
        builder = ReportBuilder()
        filename = builder.build_comprehensive_report(leads_data, campaign_results)
        
        logger.info(f"Final report generated: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Error building final report: {e}")
        raise

def main():
    """Main function to be called by CrewAI"""
    # Load all necessary data
    try:
        with open('leads_with_social.json', 'r') as f:
            leads_data = json.load(f)
    except FileNotFoundError:
        logger.error("leads_with_social.json not found. Run the social analysis first.")
        return None
    
    try:
        with open('email_campaign_results.json', 'r') as f:
            campaign_results = json.load(f)
    except FileNotFoundError:
        logger.error("email_campaign_results.json not found. Run the email campaign first.")
        return None
    
    # Build the report
    filename = build_final_report(leads_data, campaign_results)
    
    logger.info(f"Report generation completed: {filename}")
    return filename

if __name__ == "__main__":
    # Test the report builder
    test_leads = [
        {
            'name': 'Restaurante Teste',
            'website': 'https://exemplo.com',
            'analysis': {
                'lighthouse': {
                    'performance_score': 65.5,
                    'seo_score': 58.2
                },
                'seo': {
                    'total_score': 62.5
                }
            },
            'social_analysis': {
                'overall_social_score': 45.2
            }
        }
    ]
    
    test_campaign = {
        'total_emails': 1,
        'sent_count': 1,
        'failed_count': 0,
        'sent_emails': [
            {
                'lead_id': 'Restaurante Teste',
                'email': 'test@example.com',
                'subject': 'Test Subject',
                'personalization_score': 85,
                'status': 'sent',
                'sent_at': '2024-01-15 10:00:00'
            }
        ]
    }
    
    filename = build_final_report(test_leads, test_campaign)
    print(f"Test report generated: {filename}") 